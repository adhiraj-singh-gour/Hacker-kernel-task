from django.views.generic import ListView, CreateView
from .models import *
from .forms import *
from django.urls import reverse_lazy
from django.core.paginator import Paginator
import openpyxl
from openpyxl import Workbook
from django.http import HttpResponse

class AuthorCreateView(CreateView):
    model = Author
    form_class = AuthorForm
    template_name = 'Authors.html'
    success_url = reverse_lazy('authors_list')

class BookCreateView(CreateView):
    model = Book
    form_class = BookForm
    template_name = 'Book.html'
    success_url = reverse_lazy('books_list')

class BorrowRecordCreateView(CreateView):
    model = BorrowRecord
    form_class = BorrowRecordForm
    template_name = 'Borrowrecord.html'
    success_url = reverse_lazy('borrowrecords_list')


class AuthorListView(ListView):
    model = Author
    paginate_by = 10
    template_name = 'Authors_list.html'

class BookListView(ListView):
    model = Book
    paginate_by = 10
    template_name = 'Books_list.html'

class BorrowRecordListView(ListView):
    model = BorrowRecord
    paginate_by = 10
    template_name = 'Borrowrecords_list.html'


def Author_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    headers = ['Name', 'Email','Bio','ID'] 
    ws.append(headers)
    data = Author.objects.all().values_list('name', 'email','bio','ID') 
    for row in data:
        ws.append(row)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    wb.save(response)
    return response

def Books_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    headers = ['Title', 'Genre','published_date','Author','ID'] 
    ws.append(headers)
    data = Book.objects.all().values_list('title', 'genre','published_date','author','ID') 
    for row in data:
        ws.append(row)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    wb.save(response)
    return response

def Borrowrecords_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    headers = ['user_name','book','borrow_date','return_date','ID'] 
    ws.append(headers)
    data = BorrowRecord.objects.all().values_list('user_name','book','borrow_date','return_date','ID') 
    for row in data:
        ws.append(row)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    wb.save(response)
    return response





